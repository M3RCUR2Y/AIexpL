import os
import re
import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Border, Alignment


def parse_chinese_date(date_str):
    """
    解析多种日期格式，并返回对应的 datetime 对象。

    支持的格式包括：
    - 2025/9/11 8:11:11
    - 2025-09-11 08:11:22
    - 2025年9月11日
    - 2025/9/11
    - 2025-09-11

    :param date_str: 输入的日期字符串
    :return: datetime.datetime 或 None (如果无法匹配)
    """
    if not isinstance(date_str, str):
        return None

    date_str = date_str.strip()

    # 定义多种日期格式的正则表达式
    patterns = [
        # 格式: 2025/9/11 8:11:11
        (r'^(\d{4})/(\d{1,2})/(\d{1,2})\s+(\d{1,2}):(\d{1,2}):(\d{1,2})$'),
        # 格式: 2025-09-11 08:11:22
        (r'^(\d{4})-(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{1,2}):(\d{1,2})$'),
        # 格式: 2025年9月11日
        (r'^(\d{4})年(\d{1,2})月(\d{1,2})日$'),
        # 格式: 2025/9/11
        (r'^(\d{4})/(\d{1,2})/(\d{1,2})$'),
        # 格式: 2025-09-11
        (r'^(\d{4})-(\d{1,2})-(\d{1,2})$')
    ]

    # 尝试每种格式
    for pattern in patterns:
        match = re.match(pattern, date_str)
        if match:
            try:
                # 提取年月日时分秒
                groups = match.groups()
                year = int(groups[0])

                if '年' in date_str:
                    # 中文格式: YYYY年MM月DD日
                    month = int(groups[1])
                    day = int(groups[2])
                    hour = 0
                    minute = 0
                    second = 0
                    if len(groups) > 3:
                        hour = int(groups[3])
                        minute = int(groups[4])
                        second = int(groups[5])
                elif '/' in date_str or '-' in date_str:
                    # 数字格式: YYYY/MM/DD[ HH:MM:SS] 或 YYYY-MM-DD[ HH:MM:SS]
                    month = int(groups[1])
                    day = int(groups[2])
                    hour = 0
                    minute = 0
                    second = 0
                    if len(groups) > 3:
                        hour = int(groups[3])
                        minute = int(groups[4])
                        second = int(groups[5])
                else:
                    continue

                dt = datetime(year=year, month=month, day=day, hour=hour, minute=minute, second=second)
                return dt
            except ValueError:
                continue

    return None


def convert_excel_dates_inplace(file_path):
    """
    直接在原文件上处理日期格式转换，保持所有样式不变

    :param file_path: Excel 文件路径
    """
    print(f"[INFO] 正在处理文件: {file_path}")

    # 加载工作簿，保持原有格式
    wb = load_workbook(file_path)

    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  处理工作表: {sheet_name}")

        # 遍历所有单元格
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    # 尝试解析日期
                    parsed_dt = parse_chinese_date(cell.value)
                    if parsed_dt:
                        # 转换为标准格式
                        formatted_date = parsed_dt.strftime("%Y-%m-%d %H:%M:%S")
                        # 保存原有样式
                        old_font = copy.copy(cell.font)
                        old_fill = copy.copy(cell.fill)
                        old_border = copy.copy(cell.border)
                        old_alignment = copy.copy(cell.alignment)
                        old_number_format = cell.number_format

                        # 更新单元格值
                        cell.value = formatted_date

                        # 恢复原有样式
                        cell.font = old_font
                        cell.fill = old_fill
                        cell.border = old_border
                        cell.alignment = old_alignment
                        cell.number_format = old_number_format

    # 保存文件
    wb.save(file_path)
    print(f"[SUCCESS] 文件处理完成: {file_path}")


def main():
    target_dir = input("请输入要处理的目录路径：").strip()

    if not os.path.isdir(target_dir):
        print("[ERROR] 目录不存在！请检查路径后再试。")
        return

    xlsx_files = [f for f in os.listdir(target_dir) if f.endswith('.xlsx') and not f.startswith('~$')]

    if not xlsx_files:
        print("[WARN] 当前目录没有找到任何 .xlsx 文件。")
        return

    total_count = len(xlsx_files)
    success_count = 0

    for idx, filename in enumerate(xlsx_files, start=1):
        full_path = os.path.join(target_dir, filename)
        try:
            convert_excel_dates_inplace(full_path)
            success_count += 1
            print(f"[SUCCESS] 已成功处理 ({idx}/{total_count}): {filename}")
        except Exception as e:
            print(f"[FAILED] 处理失败 ({idx}/{total_count}): {filename} | 错误信息: {e}")

    print("\n==日期格式化已完成==")


if __name__ == "__main__":
    main()